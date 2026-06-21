"""
blueprints/reports.py — Report generation and export routes.
Supports filtering by client, status, and date range.
Exports as CSV or Excel (xlsx).
"""

import csv
import io
from datetime import datetime

from flask import Blueprint, render_template, request, redirect, url_for, flash, Response
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import db, Task, Client, STATUSES
from forms import ReportFilterForm

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


def _build_query(client_id, status, date_from, date_to):
    """
    Build a SQLAlchemy query for tasks based on filter parameters.
    Admins see all tasks; regular users see only their own.
    """
    query = Task.query

    if not current_user.is_admin:
        query = query.filter_by(assigned_to_id=current_user.id)

    if client_id and client_id != 0:
        query = query.filter_by(client_id=client_id)

    if status:
        query = query.filter_by(status=status)

    if date_from:
        query = query.filter(Task.created_at >= datetime.combine(date_from, datetime.min.time()))

    if date_to:
        query = query.filter(Task.created_at <= datetime.combine(date_to, datetime.max.time()))

    return query.order_by(Task.created_at.desc())


@reports_bp.route("/", methods=["GET", "POST"])
@login_required
def report():
    """Display the report page with a filterable task table."""
    form = ReportFilterForm(request.args, meta={"csrf": False})

    client_id = request.args.get("client_id", type=int, default=0)
    status = request.args.get("status", "")
    date_from_str = request.args.get("date_from", "")
    date_to_str = request.args.get("date_to", "")

    # Parse dates safely
    date_from = None
    date_to = None
    try:
        if date_from_str:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        if date_to_str:
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Invalid date format. Use YYYY-MM-DD.", "warning")

    tasks = _build_query(client_id, status, date_from, date_to).all()

    # Summary counts
    summary = {s: sum(1 for t in tasks if t.status == s) for s, _ in STATUSES}

    return render_template(
        "reports/report.html",
        form=form,
        tasks=tasks,
        summary=summary,
        client_id=client_id,
        status=status,
        date_from=date_from_str,
        date_to=date_to_str,
    )


@reports_bp.route("/export/csv")
@login_required
def export_csv():
    """Export filtered tasks as a CSV file."""
    client_id = request.args.get("client_id", type=int, default=0)
    status = request.args.get("status", "")
    date_from_str = request.args.get("date_from", "")
    date_to_str = request.args.get("date_to", "")

    date_from = _parse_date(date_from_str)
    date_to = _parse_date(date_to_str)

    tasks = _build_query(client_id, status, date_from, date_to).all()

    # Build CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "ID", "Title", "Client", "Role", "Assigned To", "Status",
        "Deadline", "Created At", "Description"
    ])

    # Data rows
    for task in tasks:
        writer.writerow([
            task.id,
            task.title,
            task.client.name,
            task.role.capitalize(),
            task.assignee.username if task.assignee else "Unassigned",
            task.status.capitalize(),
            task.deadline.strftime("%Y-%m-%d") if task.deadline else "",
            task.created_at.strftime("%Y-%m-%d %H:%M"),
            task.description or "",
        ])

    output.seek(0)
    filename = f"hitv_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@reports_bp.route("/export/excel")
@login_required
def export_excel():
    """Export filtered tasks as an Excel (.xlsx) file."""
    client_id = request.args.get("client_id", type=int, default=0)
    status = request.args.get("status", "")
    date_from_str = request.args.get("date_from", "")
    date_to_str = request.args.get("date_to", "")

    date_from = _parse_date(date_from_str)
    date_to = _parse_date(date_to_str)

    tasks = _build_query(client_id, status, date_from, date_to).all()

    # ── Build the workbook ─────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "HiTV Tasks"

    # Header style
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Title", "Client", "Role", "Assigned To", "Status",
               "Deadline", "Created At", "Description"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Status badge colors
    status_colors = {
        "pending": "FFF3CD",   # amber
        "ongoing": "CCE5FF",   # blue
        "completed": "D4EDDA", # green
    }

    # Data rows
    for row_idx, task in enumerate(tasks, start=2):
        row_data = [
            task.id,
            task.title,
            task.client.name,
            task.role.capitalize(),
            task.assignee.username if task.assignee else "Unassigned",
            task.status.capitalize(),
            task.deadline.strftime("%Y-%m-%d") if task.deadline else "",
            task.created_at.strftime("%Y-%m-%d %H:%M"),
            task.description or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Colour-code status column (col 6)
            if col_idx == 6:
                fill_color = status_colors.get(task.status, "FFFFFF")
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Write to BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"hitv_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _parse_date(date_str: str):
    """Parse a YYYY-MM-DD string into a date object, or return None."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None
